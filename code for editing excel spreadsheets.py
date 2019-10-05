import openpyxl as xl
wb = xl.load_workbook('Book1.xlsx')
sheet= wb['Sheet1']
cell = sheet.cell(1,1)

for row in range(2,sheet.max_row +1):
    cell = sheet.cell(row,3)
    mul = cell.value * 0.9
    corrected_price_cell =sheet.cell(row, 4)
    corrected_price_cell.value = mul

wb.save('test to see if exel code works.xlsx')
